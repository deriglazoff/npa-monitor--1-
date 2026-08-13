"""Экспорт выгрузки в CSV и XLSX."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import COLUMNS, HEADERS_RU, Document

WIDTHS = {
    "source": 22,
    "doc_type": 26,
    "number": 14,
    "title": 90,
    "publication_date": 16,
    "status_change_date": 18,
    "status": 46,
    "department": 34,
    "url": 46,
    "topics": 30,
    "keywords": 30,
}


def to_csv(docs: list[Document], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    # utf-8-sig — чтобы Excel открывал кириллицу без плясок с кодировкой
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNS, delimiter=";")
        writer.writerow({c: HEADERS_RU[c] for c in COLUMNS})
        for doc in docs:
            writer.writerow(doc.as_row())
    return path


def to_xlsx(docs: list[Document], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Документы"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")

    ws.append([HEADERS_RU[c] for c in COLUMNS])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for doc in docs:
        row = doc.as_row()
        ws.append([row[c] for c in COLUMNS])

    for idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = WIDTHS.get(col, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    return path
